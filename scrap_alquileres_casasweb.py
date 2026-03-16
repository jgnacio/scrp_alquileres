import cloudscraper
import requests
from bs4 import BeautifulSoup
import json
import logging
import io
import re
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# Avoid verbose logs
logging.basicConfig(level=logging.INFO)

def fetch_html(url):
    scraper = cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'desktop': True
        }
    )
    
    try:
        response = scraper.get(url, timeout=15)
        print(f"Status Code: {response.status_code}")
        if response.status_code == 200:
            return response.text
        else:
            print("Failed to get 200 OK. Content preview:")
            print(response.text[:500])
            return None
    except Exception as e:
        print(f"Exception fetching url: {e}")
        return None

def parse_html_alquileres_casasweb(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    
    # In casasweb, results are in .item-grid
    articles = soup.select(".item-grid .card")
    print(f"Se encontraron {len(articles)} avisos en el HTML de casasweb.")

    inmuebles = []

    for article in articles:
        try:
            inmueble = {}
            
            a_tag = article.select_one("a")
            if a_tag:
                href = a_tag.get("href", "")
                if href and not href.startswith("http"):
                    inmueble['url'] = "https://casasweb.com/" + href
                else:
                    inmueble['url'] = href
            else:
                inmueble['url'] = ""
                
            inmueble['id'] = ""
            if inmueble['url']:
                match = re.search(r'CW(\d+)', inmueble['url'])
                if match:
                    inmueble['id'] = match.group(1)

            # Title
            h3_title = article.select_one(".item-title h3")
            inmueble['titulo'] = h3_title.text.strip().replace("\n", " ") if h3_title else ""
            
            # Area and type
            area_tag = article.select_one(".tipo-propiedad-zona b i")
            inmueble['area_m2'] = area_tag.text.replace("m", "").replace("2", "").strip() if area_tag else ""
            
            # Details (bedrooms, condition, etc)
            det_tag = article.select_one(".item-det")
            inmueble['detalles'] = det_tag.text.strip().replace("\n", " ") if det_tag else ""
            
            # Price
            precio_container = article.select_one(".item-precio .precio")
            if precio_container:
                moneda_h2 = precio_container.select_one("h2")
                if moneda_h2:
                    moneda_small = moneda_h2.select_one("small")
                    if moneda_small:
                        inmueble['moneda'] = moneda_small.text.strip()
                        # Extract number
                        precio_text = moneda_h2.text.replace(inmueble['moneda'], "").replace("MES", "").strip()
                        inmueble['precio'] = precio_text
                    else:
                        inmueble['moneda'] = ""
                        inmueble['precio'] = moneda_h2.text.strip()
            else:
                inmueble['moneda'] = ""
                inmueble['precio'] = ""
                
            inmueble['tiene_telefono'] = False
            inmueble['whatsapp_link'] = ""
            
            # Image
            img_tag = article.select_one("img.fondo")
            if img_tag:
                style = img_tag.get("style", "")
                match = re.search(r"url\('([^']+)'\)", style)
                if match:
                    inmueble['imagen_url'] = match.group(1)
                else:
                    inmueble['imagen_url'] = ""
            else:
                inmueble['imagen_url'] = ""
                
            # Align keys with the ones in gallitoluis script
            # url, titulo, id, detalles, moneda, precio, area_m2, tiene_telefono, whatsapp_link, imagen_url
            
            inmueble_norm = {
                'url': inmueble.get('url', ''),
                'titulo': inmueble.get('titulo', ''),
                'id': inmueble.get('id', ''),
                'detalles': inmueble.get('detalles', ''),
                'moneda': inmueble.get('moneda', ''),
                'precio': inmueble.get('precio', ''),
                'area_m2': inmueble.get('area_m2', ''),
                'tiene_telefono': inmueble.get('tiene_telefono', False),
                'whatsapp_link': inmueble.get('whatsapp_link', ''),
                'imagen_url': inmueble.get('imagen_url', '')
            }

            inmuebles.append(inmueble_norm)
        except Exception as e:
            print(f"Error al procesar un aviso de casasweb: {e}")
            continue

    return inmuebles

def append_to_excel(datos, filename="alquileres.xlsx"):
    if not os.path.exists(filename):
        print(f"El archivo {filename} no existe. Por favor, corre primero el scraper de gallito.")
        return
        
    wb = load_workbook(filename)
    if "Alquileres" in wb.sheetnames:
        ws = wb["Alquileres"]
    else:
        ws = wb.active
        
    start_row = ws.max_row + 1
    
    # Read headers to know columns
    headers = [cell.value for cell in ws[1]]
    if 'imagen' in headers:
        img_idx = headers.index('imagen')
    else:
        img_idx = -1

    img_col_letter = chr(65 + img_idx) if img_idx != -1 else ""

    print("Descargando imágenes y añadiendo al Excel...")
    
    session = requests.Session()

    for i, data_dict in enumerate(datos, start=start_row): 
        # Convert dictionary data to row list based on headers order, assume keys are standard
        row_data = []
        img_url = ""
        
        # Mapping standard dict to header structure if possible, else just use standard order
        # The first scraper generates this header order: 
        # [url, titulo, id, detalles, moneda, precio, area_m2, tiene_telefono, whatsapp_link, imagen]
        for idx, key in enumerate(data_dict.keys()):
            val = data_dict[key]
            if idx == img_idx:
                img_url = val
                row_data.append("") # empty text for image cell
            else:
                row_data.append(val)
                
        ws.row_dimensions[i].height = 100
        ws.append(row_data)
        
        # Insert image
        if img_idx != -1 and img_url:
            try:
                headers_img = {
                    "User-Agent": "Mozilla/5.0"
                }
                response = session.get(img_url, headers=headers_img, timeout=10)
                if response.status_code == 200:
                    image_file = io.BytesIO(response.content)
                    img_pil = PILImage.open(image_file)
                    img_pil.thumbnail((200, 150))
                    
                    img_bytes = io.BytesIO()
                    img_pil.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    
                    img_obj = OpenpyxlImage(img_bytes)
                    cell_address = f"{img_col_letter}{i}"
                    ws.add_image(img_obj, cell_address)
            except Exception as e:
                print(f"Error descargando imagen {img_url}: {e}")
                ws[f"{img_col_letter}{i}"] = "Error img"

    wb.save(filename)
    print(f"Datos añadidos exitosamente al final de '{filename}'")

def main():
    url = "https://casasweb.com/resultados.aspx?a=Alquiler_Casa&z=0&x=1&m=&n=A&t=c&d=x&b=x&i=0&h=0&c=1#"
    html_content = fetch_html(url)
    
    if not html_content:
        print("No se pudo obtener el HTML de casasweb. Saliendo.")
        return

    datos = parse_html_alquileres_casasweb(html_content)
    
    if datos:
        print(f"\nSe han extraído {len(datos)} inmuebles de casasweb con éxito.\n")
        
        append_to_excel(datos, "alquileres.xlsx")
            
        print("\nEjemplo del primer inmueble extraído de casasweb:")
        print(json.dumps(datos[0], ensure_ascii=False, indent=4))
    else:
        print("No se extrajeron datos de casasweb.")

if __name__ == "__main__":
    main()
