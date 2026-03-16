import cloudscraper
from bs4 import BeautifulSoup
import json
import csv
import io
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage

def fetch_html(url):
    scraper = cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'desktop': True
        }
    )
    
    try:
        response = scraper.get(url, timeout=15)
        print(f"Status Code: {response.status_code}")
        if response.status_code == 200:
            return response.text
        else:
            print("Failed to get 200 OK. Content preview:")
            print(response.text[:500])
            return None
    except Exception as e:
        print(f"Exception fetching url: {e}")
        return None

def parse_html_alquileres(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    
    articles = soup.select("#grillaavisos article")
    print(f"Se encontraron {len(articles)} avisos en el HTML.")

    inmuebles = []

    for article in articles:
        try:
            inmueble = {}
            
            # Enlace y Título principal
            a_info = article.select_one(".mas-info a")
            if a_info:
                inmueble['url'] = a_info.get("href", "")
                h2 = a_info.select_one("h2")
                inmueble['titulo'] = h2.text.strip() if h2 else ""
            else:
                inmueble['url'] = ""
                inmueble['titulo'] = ""
                
            if inmueble['url']:
                inmueble['id'] = inmueble['url'].split('-')[-1]
            else:
                inmueble['id'] = ""

            # Detalles básicos
            p_details = article.select_one(".contenedor-info > div > p")
            inmueble['detalles'] = p_details.text.strip() if p_details else ""

            # Precio y Moneda
            strong_price = article.select_one(".contenedor-info > div > strong")
            if strong_price:
                span_moneda = strong_price.select_one("span")
                inmueble['moneda'] = span_moneda.text.strip() if span_moneda else ""
                
                if span_moneda:
                    precio_texto = strong_price.text.replace(span_moneda.text, "").strip()
                else:
                    precio_texto = strong_price.text.strip()
                inmueble['precio'] = precio_texto
            else:
                inmueble['moneda'] = ""
                inmueble['precio'] = ""

            # Metros cuadrados
            span_area = article.select_one(".mas-info a span")
            inmueble['area_m2'] = span_area.text.strip() if span_area else ""

            # Teléfono y Whatsapp
            telefono_tag = article.select_one(".btn-contactar i.fa-phone")
            inmueble['tiene_telefono'] = bool(telefono_tag)
            
            whatsapp_tag = article.select_one(".btn-contactar i.fa-whatsapp")
            if whatsapp_tag:
                btn_ws = whatsapp_tag.find_parent("span", class_="btn-ws-chat")
                inmueble['whatsapp_link'] = btn_ws.get("data-lnk", "") if btn_ws else ""
            else:
                inmueble['whatsapp_link'] = ""

            # Imagen principal
            img_element = article.select_one("img.img-seva")
            if not img_element:
                img_element = article.select_one(".carousel-inner .item.active img")
            
            inmueble['imagen_url'] = img_element.get("src", "") if img_element else ""

            inmuebles.append(inmueble)
        except Exception as e:
            print(f"Error al procesar un aviso: {e}")
            continue

    return inmuebles

def save_to_excel(datos, filename="alquileres.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alquileres"

    # Definir el orden deseado de las columnas
    # imagen_url se convertirá en la imagen real en la primera columna
    column_order = [
        'imagen_url', 'es_dueno', 'moneda', 'precio', 'titulo', 'detalles', 
        'area_m2', 'id', 'tiene_telefono', 'whatsapp_link', 'url'
    ]
    
    # Nombres amigables para los encabezados
    header_labels = {
        'imagen_url': 'Imagen',
        'es_dueno': '¿Dueño?',
        'moneda': 'Moneda',
        'precio': 'Precio',
        'titulo': 'Título',
        'detalles': 'Detalles',
        'area_m2': 'Área (m2)',
        'id': 'ID',
        'tiene_telefono': 'Teléfono',
        'whatsapp_link': 'WhatsApp',
        'url': 'Enlace'
    }

    # Escribir encabezados con estilo
    headers = [header_labels.get(col, col) for col in column_order]
    ws.append(headers)
    
    # Estilo para el encabezado
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned

    # Configurar anchos de columna
    col_widths = {
        'A': 30,  # Imagen
        'B': 10,  # Dueño
        'C': 10,  # Moneda
        'D': 15,  # Precio
        'E': 40,  # Título
        'F': 60,  # Detalles
        'G': 12,  # Área
        'H': 12,  # ID
        'I': 10,  # Teléfono
        'J': 12,  # WhatsApp
        'K': 30   # Enlace
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    print("Descargando imágenes y generando Excel con formato premium...")
    
    session = requests.Session()
    headers_img = {"User-Agent": "Mozilla/5.0"}

    # Bordes para las celdas
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )

    for i, inmueble in enumerate(datos, start=2): 
        # Extraer datos en el orden correcto
        row_data = []
        for col in column_order:
            if col == 'imagen_url':
                row_data.append("") # Dejar espacio para la imagen
            else:
                row_data.append(inmueble.get(col, ""))
        
        ws.append(row_data)
        ws.row_dimensions[i].height = 110 # Altura para la miniatura
        
        # Aplicar estilos a las celdas de la fila
        for col_index, cell in enumerate(ws[i]):
            cell.alignment = Alignment(vertical="center", wrap_text=(cell.column_letter == 'F' or cell.column_letter == 'E'))
            cell.border = thin_border
            # Centrar algunas columnas específicas
            if cell.column_letter in ['A', 'B', 'C', 'D', 'G', 'H', 'I', 'J']:
                 cell.alignment = center_aligned

            # Traducir booleano de dueño a SI/NO
            if column_order[col_index] == 'es_dueno':
                if cell.value:
                    cell.value = "SÍ"
                    cell.font = Font(bold=True, color="27AE60") # Verde para destacados
                else:
                    cell.value = ""

            # Hacer que el enlace sea clickeable
            if column_order[col_index] == 'url' and cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

        # Insertar imagen
        img_url = inmueble.get('imagen_url', '')
        if img_url:
            try:
                response = session.get(img_url, headers=headers_img, timeout=10)
                if response.status_code == 200:
                    image_file = io.BytesIO(response.content)
                    img_pil = PILImage.open(image_file)
                    # Redimensionar para que quepa bien
                    img_pil.thumbnail((200, 140))
                    
                    img_bytes = io.BytesIO()
                    img_pil.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    
                    img_obj = OpenpyxlImage(img_bytes)
                    img_obj.anchor = f"A{i}"
                    ws.add_image(img_obj)
            except Exception as e:
                print(f"Error descargando imagen {img_url}: {e}")
                ws[f"A{i}"] = "Sin imagen"
                ws[f"A{i}"].font = Font(italic=True, color="95A5A6")

    # Activar el filtro automático
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"Excel premium generado exitosamente en '{filename}'")

def main():
    url = "https://www.gallito.com.uy/inmuebles/casas/alquiler/montevideo/otros/pre-0-21000-pesos"
    html_content = fetch_html(url)
    
    if not html_content:
        print("No se pudo obtener el HTML. Saliendo.")
        return

    datos = parse_html_alquileres(html_content)
    
    if datos:
        print(f"\nSe han extraído {len(datos)} inmuebles con éxito.\n")
        
        save_to_excel(datos, "alquileres.xlsx")
            
        print("\nEjemplo del primer inmueble extraído:")
        print(json.dumps(datos[0], ensure_ascii=False, indent=4))
    else:
        print("No se extrajeron datos.")

if __name__ == "__main__":
    main()
